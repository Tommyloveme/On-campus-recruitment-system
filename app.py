# -*- coding: utf-8 -*-
"""校招入职跟踪管理系统 - 跨平台（Windows / SUSE Linux）

面向已接收 Offer 的校招候选人，跟踪签约、体检、入职预约到正式入职的全流程。
技术栈：Flask + SQLite（标准库 sqlite3），Excel 导入使用 openpyxl。
字段通过 config/fields.json 配置：导入哪些列、网页显示哪些列均可配置。
"""
import io
import json
import logging
import os
import re
import sqlite3
import sys
import secrets
import threading
import time
import zipfile
from datetime import datetime
from functools import wraps

from flask import Flask, g, jsonify, request, session, send_file, send_from_directory
from werkzeug.security import check_password_hash, generate_password_hash
from openpyxl import Workbook, load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "data", "candidates.db")
CONFIG_PATH = os.path.join(BASE_DIR, "config", "fields.json")
APP_CONFIG_PATH = os.path.join(BASE_DIR, "config", "app_config.json")
SECRET_PATH = os.path.join(BASE_DIR, "data", ".secret_key")
RESUME_DIR = os.path.join(BASE_DIR, "data", "resumes")
BACKUP_DIR = os.path.join(BASE_DIR, "data", "backups")


def load_app_config():
    """运行配置（config/app_config.json）：端口、线程数、分页大小、默认密码等易变内容。"""
    with open(APP_CONFIG_PATH, encoding="utf-8") as f:
        return json.load(f)


APP_CONFIG = load_app_config()
ALLOWED_RESUME_EXT = set(APP_CONFIG["resume"]["allowed_ext"])


def setup_logging():
    """运行日志写入 data/server.log 并同步输出控制台；级别由 LOG_LEVEL 环境变量或 app_config 控制。"""
    level_name = os.environ.get("LOG_LEVEL") or APP_CONFIG.get("logging", {}).get("level", "INFO")
    level = getattr(logging, level_name.upper(), logging.INFO)
    os.makedirs(os.path.join(BASE_DIR, "data"), exist_ok=True)
    fmt = logging.Formatter("%(asctime)s %(levelname)s [%(name)s] %(message)s", datefmt="%H:%M:%S")
    root = logging.getLogger("campus")
    root.setLevel(level)
    root.propagate = False
    if not root.handlers:
        fh = logging.FileHandler(os.path.join(BASE_DIR, "data", "server.log"), encoding="utf-8")
        fh.setFormatter(fmt)
        root.addHandler(fh)
        # 交互式启动时同步输出控制台；脚本后台启动时 stderr 已重定向至 server.log，不再重复添加
        if sys.stderr.isatty():
            sh = logging.StreamHandler()
            sh.setFormatter(fmt)
            root.addHandler(sh)
    return root


log = setup_logging()


def _who(user):
    """日志中标识操作用户（用户名/显示名/角色）。"""
    return f"{user['username']}({user['display_name']}/{user['role']})"

app = Flask(__name__, static_folder="static", static_url_path="/static")
app.config["MAX_CONTENT_LENGTH"] = APP_CONFIG["server"]["max_upload_mb"] * 1024 * 1024


# ---------------------------------------------------------------- 基础设施
def _ensure_secret():
    os.makedirs(os.path.dirname(SECRET_PATH), exist_ok=True)
    if not os.path.exists(SECRET_PATH):
        with open(SECRET_PATH, "w", encoding="utf-8") as f:
            f.write(secrets.token_hex(32))
    with open(SECRET_PATH, encoding="utf-8") as f:
        return f.read().strip()


app.secret_key = _ensure_secret()


def get_db():
    if "db" not in g:
        # 多线程并发：每个请求独立连接 + WAL 模式 + 写锁等待，支撑数百人同时使用
        g.db = sqlite3.connect(DB_PATH, timeout=15)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
        g.db.execute("PRAGMA busy_timeout = 10000")
    return g.db


@app.teardown_appcontext
def close_db(_exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def load_fields(group_id=None):
    """主配置定义全部字段；每个分组可有独立的显示配置文件覆盖 visible 开关。"""
    with open(CONFIG_PATH, encoding="utf-8") as f:
        fields = json.load(f)["fields"]
    if group_id:
        path = group_config_path(group_id)
        if os.path.exists(path):
            with open(path, encoding="utf-8") as f:
                visible_map = json.load(f).get("visible", {})
            for field in fields:
                if field["key"] in visible_map:
                    field["visible"] = bool(visible_map[field["key"]])
    return fields


def group_config_path(group_id):
    return os.path.join(BASE_DIR, "config", f"fields_group_{group_id}.json")


def save_fields(fields):
    with open(CONFIG_PATH, encoding="utf-8") as f:
        cfg = json.load(f)
    cfg["fields"] = fields
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def save_group_visible(group_id, visible_map):
    path = group_config_path(group_id)
    data = {"comment": "该分组的网页字段显示配置，覆盖主配置 fields.json 中的 visible 开关",
            "visible": visible_map}
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ---------------------------------------------------------------- 数据库初始化
SCHEMA = """
CREATE TABLE IF NOT EXISTS groups (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT UNIQUE NOT NULL,
    created_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT UNIQUE NOT NULL,
    display_name TEXT NOT NULL,
    password_hash TEXT NOT NULL,
    role TEXT NOT NULL DEFAULT 'editor',   -- admin / editor / viewer
    group_id INTEGER REFERENCES groups(id),
    created_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS candidates (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    group_id INTEGER NOT NULL REFERENCES groups(id),
    data TEXT NOT NULL,                    -- JSON: {field_key: value}
    resume_file TEXT,                      -- 简历在磁盘上的存储文件名
    resume_name TEXT,                      -- 简历原始文件名
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS logs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_name TEXT NOT NULL,
    action TEXT NOT NULL,                  -- create / update / delete / import
    candidate_id INTEGER,
    candidate_name TEXT,
    group_id INTEGER,
    message TEXT NOT NULL,
    created_at TEXT NOT NULL
);
"""


def _migrate(db):
    """为老数据库补充新列。"""
    cols = {r["name"] for r in db.execute("PRAGMA table_info(candidates)").fetchall()}
    if "resume_file" not in cols:
        db.execute("ALTER TABLE candidates ADD COLUMN resume_file TEXT")
        db.execute("ALTER TABLE candidates ADD COLUMN resume_name TEXT")
        db.commit()


def init_db(demo=False):
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    os.makedirs(RESUME_DIR, exist_ok=True)
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA journal_mode = WAL")   # 读写不互斥，大幅提升并发能力
    db.executescript(SCHEMA)
    _migrate(db)
    cur = db.execute("SELECT COUNT(*) AS c FROM users")
    if cur.fetchone()["c"] == 0:
        db.execute(
            "INSERT INTO users (username, display_name, password_hash, role, group_id, created_at) "
            "VALUES (?,?,?,?,?,?)",
            ("admin", "系统管理员", generate_password_hash("admin123"), "admin", None, now_str()),
        )
        db.commit()
        print("已创建默认管理员账号: admin / admin123")
    if demo:
        _seed_demo(db)
    db.close()


def _seed_demo(db):
    if db.execute("SELECT COUNT(*) AS c FROM groups").fetchone()["c"] > 0:
        print("已存在分组数据，跳过示例数据。")
        return
    g1 = db.execute("INSERT INTO groups (name, created_at) VALUES (?,?)", ("研发一组", now_str())).lastrowid
    g2 = db.execute("INSERT INTO groups (name, created_at) VALUES (?,?)", ("研发二组", now_str())).lastrowid
    db.execute(
        "INSERT INTO users (username, display_name, password_hash, role, group_id, created_at) VALUES (?,?,?,?,?,?)",
        ("lead01", "组管理员-老张", generate_password_hash("123456"), "group_admin", g1, now_str()),
    )
    db.execute(
        "INSERT INTO users (username, display_name, password_hash, role, group_id, created_at) VALUES (?,?,?,?,?,?)",
        ("hr01", "招聘专员-小王", generate_password_hash("123456"), "editor", g1, now_str()),
    )
    db.execute(
        "INSERT INTO users (username, display_name, password_hash, role, group_id, created_at) VALUES (?,?,?,?,?,?)",
        ("hr02", "招聘专员-小李", generate_password_hash("123456"), "editor", g2, now_str()),
    )
    samples = [
        (g1, {"name": "张伟", "phone": "13800000001", "interface_manager": "王经理", "interface_person": "刘洋",
              "dept_level2": "云计算BG", "dept_level3": "存储部", "offer_status": "已接受", "sign_status": "已签约",
              "work_location": "深圳", "graduation_time": "2026-06-30", "expected_onboard_time": "2026-07-15",
              "physical_exam_time": "2026-06-20", "physical_exam_done": "否", "onboard_booked": "是",
              "onboard_booked_time": "2026-07-15", "onboarded": "否", "onboard_risk": "低"}),
        (g1, {"name": "李娜", "phone": "13800000002", "interface_manager": "王经理", "interface_person": "刘洋",
              "dept_level2": "云计算BG", "dept_level3": "计算部", "offer_status": "已发放", "sign_status": "未签约",
              "work_location": "杭州", "graduation_time": "2026-06-30", "expected_onboard_time": "2026-08-01",
              "physical_exam_done": "否", "onboard_booked": "否", "onboarded": "否", "onboard_risk": "中"}),
        (g2, {"name": "陈强", "phone": "13800000003", "interface_manager": "赵经理", "interface_person": "孙敏",
              "dept_level2": "终端BG", "dept_level3": "软件部", "offer_status": "已接受", "sign_status": "已签约",
              "work_location": "上海", "graduation_time": "2026-07-01", "expected_onboard_time": "2026-07-20",
              "physical_exam_time": "2026-06-25", "physical_exam_done": "是", "onboard_booked": "是",
              "onboard_booked_time": "2026-07-20", "onboarded": "否", "onboard_risk": "无"}),
    ]
    for gid, data in samples:
        db.execute(
            "INSERT INTO candidates (group_id, data, created_at, updated_at) VALUES (?,?,?,?)",
            (gid, json.dumps(data, ensure_ascii=False), now_str(), now_str()),
        )
    db.commit()
    print("已写入示例分组/用户/候选人数据（lead01 组管理员、hr01、hr02 / 123456）。")


# ---------------------------------------------------------------- 权限
def current_user():
    uid = session.get("uid")
    if not uid:
        return None
    return get_db().execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()


def login_required(fn):
    @wraps(fn)
    def wrapper(*a, **kw):
        user = current_user()
        if not user:
            return jsonify({"error": "未登录"}), 401
        g.user = user
        return fn(*a, **kw)
    return wrapper


def admin_required(fn):
    @wraps(fn)
    def wrapper(*a, **kw):
        user = current_user()
        if not user:
            return jsonify({"error": "未登录"}), 401
        if user["role"] != "admin":
            return jsonify({"error": "需要管理员权限"}), 403
        g.user = user
        return fn(*a, **kw)
    return wrapper


# 角色说明：
#   admin         系统管理员：全部权限
#   global_viewer 全局查看员：界面与系统管理员一致（全分组候选人/总览/图表/日志），但无系统管理入口，所有数据只读
#   group_admin   组管理员：本组候选人增/删/改/查 + 添加本组成员
#   editor        组成员（同学）：本组候选人增/改/查，无删除权限
#   viewer        只读：仅查看本组
VALID_ROLES = ("admin", "global_viewer", "group_admin", "editor", "viewer")
GLOBAL_VIEW_ROLES = ("admin", "global_viewer")   # 可跨分组查看数据的角色


def can_edit_group(user, group_id):
    if user["role"] == "admin":
        return True
    return user["role"] in ("group_admin", "editor") and user["group_id"] == group_id


def can_delete_group(user, group_id):
    if user["role"] == "admin":
        return True
    return user["role"] == "group_admin" and user["group_id"] == group_id


def can_view_group(user, group_id):
    return user["role"] in GLOBAL_VIEW_ROLES or user["group_id"] == group_id


# ---------------------------------------------------------------- 日志
def add_log(user, action, message, candidate_id=None, candidate_name=None, group_id=None):
    get_db().execute(
        "INSERT INTO logs (user_name, action, candidate_id, candidate_name, group_id, message, created_at) "
        "VALUES (?,?,?,?,?,?,?)",
        (user["display_name"], action, candidate_id, candidate_name, group_id, message, now_str()),
    )


# ---------------------------------------------------------------- 认证接口
@app.post("/api/login")
def api_login():
    body = request.get_json(force=True)
    username = body.get("username", "")
    user = get_db().execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
    if not user or not check_password_hash(user["password_hash"], body.get("password", "")):
        log.warning("登录失败 username=%s ip=%s", username, request.remote_addr)
        return jsonify({"error": "用户名或密码错误"}), 401
    session["uid"] = user["id"]
    log.info("登录成功 %s ip=%s", _who(user), request.remote_addr)
    return jsonify(user_dict(user))


@app.post("/api/logout")
def api_logout():
    uid = session.get("uid")
    if uid:
        u = get_db().execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
        if u:
            log.info("退出登录 %s", _who(u))
    session.clear()
    return jsonify({"ok": True})


@app.get("/api/me")
@login_required
def api_me():
    return jsonify(user_dict(g.user))


def user_dict(u):
    group_name = None
    if u["group_id"]:
        row = get_db().execute("SELECT name FROM groups WHERE id=?", (u["group_id"],)).fetchone()
        group_name = row["name"] if row else None
    return {"id": u["id"], "username": u["username"], "display_name": u["display_name"],
            "role": u["role"], "group_id": u["group_id"], "group_name": group_name}


# ---------------------------------------------------------------- 字段配置
@app.get("/api/config")
@login_required
def api_config():
    """普通用户自动取本组配置；管理员/全局查看员可用 ?group_id= 查看某组配置，默认全局主配置。"""
    if g.user["role"] in GLOBAL_VIEW_ROLES:
        group_id = request.args.get("group_id", type=int)
    else:
        group_id = g.user["group_id"]
    # app 部分（分页大小等界面配置）一并下发给前端
    return jsonify({"fields": load_fields(group_id), "group_id": group_id,
                    "app": load_app_config().get("ui", {})})


@app.put("/api/config")
@login_required
def api_config_update():
    """管理员：可改全局主配置或任意分组配置；组管理员：只能改本组的显示配置。"""
    body = request.get_json(force=True)
    updates = {f["key"]: f for f in body.get("fields", [])}

    if g.user["role"] == "admin":
        group_id = body.get("group_id")
    elif g.user["role"] == "group_admin":
        group_id = g.user["group_id"]
        if not group_id:
            return jsonify({"error": "您未归属任何分组"}), 400
    else:
        return jsonify({"error": "无配置权限"}), 403

    log.debug("字段配置更新 %s group_id=%s keys=%s", _who(g.user), group_id, list(updates.keys()))
    if group_id:
        # 分组级配置：仅覆盖 visible 开关，独立文件存储
        path = group_config_path(group_id)
        visible_map = {}
        if os.path.exists(path):
            with open(path, encoding="utf-8") as f:
                visible_map = json.load(f).get("visible", {})
        for key, u in updates.items():
            if "visible" in u:
                visible_map[key] = bool(u["visible"])
        save_group_visible(group_id, visible_map)
        gname = group_name_map().get(group_id, "")
        add_log(g.user, "config", f"{g.user['display_name']} 调整了「{gname}」的字段显示配置", group_id=group_id)
    else:
        fields = load_fields()
        for f in fields:
            if f["key"] in updates:
                u = updates[f["key"]]
                for attr in ("visible", "editable", "importable", "label", "excel_column"):
                    if attr in u:
                        f[attr] = u[attr]
        save_fields(fields)
        add_log(g.user, "config", f"{g.user['display_name']} 调整了全局字段显示配置")
    get_db().commit()
    log.info("字段配置已保存 %s scope=%s", _who(g.user), group_id or "global")
    return jsonify({"fields": load_fields(group_id), "group_id": group_id})


# ---------------------------------------------------------------- 分组
@app.get("/api/groups")
@login_required
def api_groups():
    rows = get_db().execute(
        "SELECT g.*, (SELECT COUNT(*) FROM candidates c WHERE c.group_id=g.id) AS candidate_count "
        "FROM groups g ORDER BY g.id"
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/groups")
@admin_required
def api_group_create():
    name = (request.get_json(force=True).get("name") or "").strip()
    if not name:
        return jsonify({"error": "分组名不能为空"}), 400
    db = get_db()
    try:
        db.execute("INSERT INTO groups (name, created_at) VALUES (?,?)", (name, now_str()))
    except sqlite3.IntegrityError:
        return jsonify({"error": "分组名已存在"}), 400
    add_log(g.user, "group", f"{g.user['display_name']} 创建了分组「{name}」")
    db.commit()
    return jsonify({"ok": True})


@app.delete("/api/groups/<int:gid>")
@admin_required
def api_group_delete(gid):
    db = get_db()
    row = db.execute("SELECT name FROM groups WHERE id=?", (gid,)).fetchone()
    if not row:
        return jsonify({"error": "分组不存在"}), 404
    if db.execute("SELECT COUNT(*) AS c FROM candidates WHERE group_id=?", (gid,)).fetchone()["c"] > 0:
        return jsonify({"error": "该分组下仍有候选人，无法删除"}), 400
    db.execute("UPDATE users SET group_id=NULL WHERE group_id=?", (gid,))
    db.execute("DELETE FROM groups WHERE id=?", (gid,))
    if os.path.exists(group_config_path(gid)):
        os.remove(group_config_path(gid))
    add_log(g.user, "group", f"{g.user['display_name']} 删除了分组「{row['name']}」")
    db.commit()
    return jsonify({"ok": True})


# ---------------------------------------------------------------- 用户管理
@app.get("/api/users")
@login_required
def api_users():
    sql = ("SELECT u.id, u.username, u.display_name, u.role, u.group_id, g.name AS group_name "
           "FROM users u LEFT JOIN groups g ON g.id=u.group_id")
    params = []
    if g.user["role"] == "admin":
        pass
    elif g.user["role"] == "group_admin":
        sql += " WHERE u.group_id=?"
        params.append(g.user["group_id"])
    else:
        return jsonify({"error": "无用户管理权限"}), 403
    rows = get_db().execute(sql + " ORDER BY u.id", params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/users")
@login_required
def api_user_create():
    b = request.get_json(force=True)
    username = (b.get("username") or "").strip()
    if not username:
        return jsonify({"error": "用户名不能为空"}), 400
    password = b.get("password") or APP_CONFIG["security"]["default_password"]  # 未填写时使用默认密码
    role = b.get("role", "editor")
    group_id = b.get("group_id")

    if g.user["role"] == "admin":
        if role not in VALID_ROLES:
            return jsonify({"error": "角色不合法"}), 400
    elif g.user["role"] == "group_admin":
        # 组管理员只能向本组添加普通成员
        if role not in ("editor", "viewer"):
            return jsonify({"error": "组管理员只能添加组成员或只读账号"}), 403
        group_id = g.user["group_id"]
    else:
        return jsonify({"error": "无添加用户权限"}), 403

    db = get_db()
    try:
        db.execute(
            "INSERT INTO users (username, display_name, password_hash, role, group_id, created_at) VALUES (?,?,?,?,?,?)",
            (username, b.get("display_name") or username, generate_password_hash(password),
             role, group_id, now_str()),
        )
    except sqlite3.IntegrityError:
        return jsonify({"error": "用户名已存在"}), 400
    add_log(g.user, "user", f"{g.user['display_name']} 创建了用户「{b.get('display_name') or username}」",
            group_id=group_id)
    db.commit()
    return jsonify({"ok": True})


@app.put("/api/users/<int:uid>")
@admin_required
def api_user_update(uid):
    b = request.get_json(force=True)
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not user:
        return jsonify({"error": "用户不存在"}), 404
    role = b.get("role", user["role"])
    if role not in VALID_ROLES:
        return jsonify({"error": "角色不合法"}), 400
    db.execute(
        "UPDATE users SET display_name=?, role=?, group_id=? WHERE id=?",
        (b.get("display_name", user["display_name"]), role, b.get("group_id", user["group_id"]), uid),
    )
    if b.get("password"):
        db.execute("UPDATE users SET password_hash=? WHERE id=?", (generate_password_hash(b["password"]), uid))
    add_log(g.user, "user", f"{g.user['display_name']} 更新了用户「{user['display_name']}」的信息")
    db.commit()
    return jsonify({"ok": True})


@app.delete("/api/users/<int:uid>")
@admin_required
def api_user_delete(uid):
    if uid == g.user["id"]:
        return jsonify({"error": "不能删除自己"}), 400
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not user:
        return jsonify({"error": "用户不存在"}), 404
    db.execute("DELETE FROM users WHERE id=?", (uid,))
    add_log(g.user, "user", f"{g.user['display_name']} 删除了用户「{user['display_name']}」")
    db.commit()
    return jsonify({"ok": True})


# ---------------------------------------------------------------- 候选人
def candidate_dict(row, group_names=None):
    data = json.loads(row["data"])
    gname = (group_names or {}).get(row["group_id"])
    if gname is None:
        r = get_db().execute("SELECT name FROM groups WHERE id=?", (row["group_id"],)).fetchone()
        gname = r["name"] if r else ""
    return {"id": row["id"], "group_id": row["group_id"], "group_name": gname,
            "updated_at": row["updated_at"], "resume_name": row["resume_name"], "data": data}


def group_name_map():
    return {r["id"]: r["name"] for r in get_db().execute("SELECT id, name FROM groups").fetchall()}


@app.get("/api/candidates")
@login_required
def api_candidates():
    db = get_db()
    sql = "SELECT * FROM candidates"
    params = []
    if g.user["role"] not in GLOBAL_VIEW_ROLES:
        if not g.user["group_id"]:
            return jsonify([])
        sql += " WHERE group_id=?"
        params.append(g.user["group_id"])
    elif request.args.get("group_id"):
        sql += " WHERE group_id=?"
        params.append(int(request.args["group_id"]))
    sql += " ORDER BY updated_at DESC"
    rows = db.execute(sql, params).fetchall()
    names = group_name_map()
    result = [candidate_dict(r, names) for r in rows]
    q = (request.args.get("q") or "").strip()
    if q:
        result = [c for c in result
                  if any(q in str(v) for v in c["data"].values()) or q in c["group_name"]]
    log.debug("候选人列表 %s 返回%d条 q=%s group=%s",
              _who(g.user), len(result), q or "-", request.args.get("group_id", "-"))
    return jsonify(result)


@app.post("/api/candidates")
@login_required
def api_candidate_create():
    b = request.get_json(force=True)
    group_id = b.get("group_id") or g.user["group_id"]
    if not group_id:
        return jsonify({"error": "请指定分组"}), 400
    if not can_edit_group(g.user, group_id):
        log.warning("新增候选人权限拒绝 %s group_id=%s", _who(g.user), group_id)
        return jsonify({"error": "无该分组的编辑权限"}), 403
    fields = load_fields()
    data = {f["key"]: str(b.get("data", {}).get(f["key"], "") or "").strip() for f in fields if f.get("editable")}
    if not data.get("name"):
        return jsonify({"error": "候选人姓名不能为空"}), 400
    db = get_db()
    cur = db.execute(
        "INSERT INTO candidates (group_id, data, created_at, updated_at) VALUES (?,?,?,?)",
        (group_id, json.dumps(data, ensure_ascii=False), now_str(), now_str()),
    )
    gname = group_name_map().get(group_id, "")
    add_log(g.user, "create", f"{g.user['display_name']} 新增了候选人「{data['name']}」（{gname}）",
            cur.lastrowid, data["name"], group_id)
    db.commit()
    log.info("新增候选人 %s id=%d name=%s group=%s", _who(g.user), cur.lastrowid, data["name"], gname)
    return jsonify({"ok": True, "id": cur.lastrowid})


@app.put("/api/candidates/<int:cid>")
@login_required
def api_candidate_update(cid):
    db = get_db()
    row = db.execute("SELECT * FROM candidates WHERE id=?", (cid,)).fetchone()
    if not row:
        return jsonify({"error": "候选人不存在"}), 404
    if not can_edit_group(g.user, row["group_id"]):
        log.warning("修改候选人权限拒绝 %s cid=%d", _who(g.user), cid)
        return jsonify({"error": "无该分组的编辑权限"}), 403
    old = json.loads(row["data"])
    incoming = request.get_json(force=True).get("data", {})
    fields = load_fields()
    labels = {f["key"]: f["label"] for f in fields}
    new = dict(old)
    changes = []
    for f in fields:
        if not f.get("editable") or f["key"] not in incoming:
            continue
        k = f["key"]
        nv = str(incoming[k] or "").strip()
        ov = str(old.get(k, "") or "")
        if nv != ov:
            new[k] = nv
            changes.append(f"{labels[k]}：{ov or '空'} → {nv or '空'}")
    if not changes:
        return jsonify({"ok": True, "changed": 0})
    if not new.get("name"):
        return jsonify({"error": "候选人姓名不能为空"}), 400
    db.execute("UPDATE candidates SET data=?, updated_at=? WHERE id=?",
               (json.dumps(new, ensure_ascii=False), now_str(), cid))
    name = new.get("name") or old.get("name", "")
    add_log(g.user, "update",
            f"{g.user['display_name']} 修改了「{name}」：" + "；".join(changes),
            cid, name, row["group_id"])
    db.commit()
    log.info("修改候选人 %s id=%d name=%s 变更%d项", _who(g.user), cid, name, len(changes))
    log.debug("修改明细 id=%d %s", cid, "；".join(changes))
    return jsonify({"ok": True, "changed": len(changes)})


@app.delete("/api/candidates/<int:cid>")
@login_required
def api_candidate_delete(cid):
    db = get_db()
    row = db.execute("SELECT * FROM candidates WHERE id=?", (cid,)).fetchone()
    if not row:
        return jsonify({"error": "候选人不存在"}), 404
    if not can_delete_group(g.user, row["group_id"]):
        log.warning("删除候选人权限拒绝 %s cid=%d", _who(g.user), cid)
        return jsonify({"error": "无删除权限（仅系统管理员和组管理员可删除）"}), 403
    name = json.loads(row["data"]).get("name", "")
    _remove_resume_file(row["resume_file"])
    db.execute("DELETE FROM candidates WHERE id=?", (cid,))
    add_log(g.user, "delete", f"{g.user['display_name']} 删除了候选人「{name}」", cid, name, row["group_id"])
    db.commit()
    log.info("删除候选人 %s id=%d name=%s", _who(g.user), cid, name)
    return jsonify({"ok": True})


@app.post("/api/candidates/batch_delete")
@login_required
def api_candidates_batch_delete():
    """批量删除：仅系统管理员（任意分组）与组管理员（本组）可执行。"""
    if g.user["role"] not in ("admin", "group_admin"):
        log.warning("批量删除权限拒绝 %s", _who(g.user))
        return jsonify({"error": "仅系统管理员和组管理员可批量删除"}), 403
    ids = request.get_json(force=True).get("ids") or []
    if not ids:
        return jsonify({"error": "请先勾选候选人"}), 400
    db = get_db()
    placeholders = ",".join("?" * len(ids))
    rows = db.execute(f"SELECT * FROM candidates WHERE id IN ({placeholders})", ids).fetchall()

    deleted_names = []
    for row in rows:
        if not can_delete_group(g.user, row["group_id"]):
            continue
        _remove_resume_file(row["resume_file"])
        db.execute("DELETE FROM candidates WHERE id=?", (row["id"],))
        deleted_names.append(json.loads(row["data"]).get("name", ""))
    if not deleted_names:
        return jsonify({"error": "选中的候选人均无删除权限"}), 403

    shown = "、".join(deleted_names[:5]) + ("等" if len(deleted_names) > 5 else "")
    add_log(g.user, "delete",
            f"{g.user['display_name']} 批量删除了 {len(deleted_names)} 名候选人（{shown}）",
            group_id=g.user["group_id"])
    db.commit()
    log.info("批量删除 %s 删除%d 跳过%d", _who(g.user), len(deleted_names), len(rows) - len(deleted_names))
    return jsonify({"ok": True, "deleted": len(deleted_names),
                    "skipped": len(rows) - len(deleted_names)})


# ---------------------------------------------------------------- 简历管理
def _remove_resume_file(stored_name):
    if not stored_name:
        return
    path = os.path.join(RESUME_DIR, stored_name)
    if os.path.exists(path):
        os.remove(path)


def _get_candidate_or_403(cid, need="edit"):
    """返回 (row, error_response)。need: view / edit / delete"""
    row = get_db().execute("SELECT * FROM candidates WHERE id=?", (cid,)).fetchone()
    if not row:
        return None, (jsonify({"error": "候选人不存在"}), 404)
    check = {"view": can_view_group, "edit": can_edit_group, "delete": can_delete_group}[need]
    if not check(g.user, row["group_id"]):
        return None, (jsonify({"error": "无该分组的操作权限"}), 403)
    return row, None


@app.post("/api/candidates/<int:cid>/resume")
@login_required
def api_resume_upload(cid):
    row, err = _get_candidate_or_403(cid)
    if err:
        return err
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error": "请选择简历文件"}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ALLOWED_RESUME_EXT:
        return jsonify({"error": "仅支持 .pdf 和 .docx 格式的简历"}), 400
    os.makedirs(RESUME_DIR, exist_ok=True)
    stored = f"{cid}_{secrets.token_hex(8)}{ext}"
    f.save(os.path.join(RESUME_DIR, stored))
    _remove_resume_file(row["resume_file"])

    db = get_db()
    db.execute("UPDATE candidates SET resume_file=?, resume_name=?, updated_at=? WHERE id=?",
               (stored, f.filename, now_str(), cid))
    name = json.loads(row["data"]).get("name", "")
    verb = "更新" if row["resume_file"] else "上传"
    add_log(g.user, "update", f"{g.user['display_name']} {verb}了「{name}」的简历（{f.filename}）",
            cid, name, row["group_id"])
    db.commit()
    log.info("简历%s %s cid=%d name=%s file=%s", verb, _who(g.user), cid, name, f.filename)
    return jsonify({"ok": True, "resume_name": f.filename})


@app.get("/api/candidates/<int:cid>/resume")
@login_required
def api_resume_download(cid):
    row, err = _get_candidate_or_403(cid, need="view")
    if err:
        return err
    if not row["resume_file"]:
        return jsonify({"error": "该候选人尚未上传简历"}), 404
    path = os.path.join(RESUME_DIR, row["resume_file"])
    if not os.path.exists(path):
        return jsonify({"error": "简历文件丢失，请重新上传"}), 404
    name = json.loads(row["data"]).get("name", "")
    ext = os.path.splitext(row["resume_name"])[1]
    return send_file(path, as_attachment=True, download_name=f"{name}_{os.path.splitext(row['resume_name'])[0]}{ext}")


PREVIEW_PAGE = """<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8">
<title>{title} - 简历预览</title>
<style>
body {{ font-family:"Segoe UI","Microsoft YaHei",sans-serif; background:#f1f5f9; margin:0; }}
.page {{ max-width:860px; margin:24px auto; background:#fff; border-radius:12px;
        padding:40px 48px; box-shadow:0 4px 24px rgba(0,0,0,.08); line-height:1.8; color:#0f172a; }}
.page img {{ max-width:100%; }}
h1,h2,h3 {{ color:#1d4ed8; }}
table {{ border-collapse:collapse; }} td,th {{ border:1px solid #e2e8f0; padding:4px 10px; }}
.tip {{ text-align:center; color:#64748b; font-size:12px; margin:12px 0 24px; }}
</style></head><body>
<div class="tip">简历预览：{title}（如格式有出入，请下载原文件查看）</div>
<div class="page">{content}</div>
</body></html>"""


@app.get("/api/candidates/<int:cid>/resume/preview")
@login_required
def api_resume_preview(cid):
    row, err = _get_candidate_or_403(cid, need="view")
    if err:
        return err
    if not row["resume_file"]:
        return jsonify({"error": "该候选人尚未上传简历"}), 404
    path = os.path.join(RESUME_DIR, row["resume_file"])
    if not os.path.exists(path):
        return jsonify({"error": "简历文件丢失，请重新上传"}), 404

    ext = os.path.splitext(row["resume_file"])[1].lower()
    if ext == ".pdf":
        # 浏览器内置 PDF 阅读器在线查看
        return send_file(path, mimetype="application/pdf", as_attachment=False,
                         download_name=row["resume_name"])
    # docx -> HTML
    try:
        import mammoth
        with open(path, "rb") as f:
            html = mammoth.convert_to_html(f).value
    except Exception:
        return jsonify({"error": "简历解析失败，请下载原文件查看"}), 500
    name = json.loads(row["data"]).get("name", "")
    return PREVIEW_PAGE.format(title=f"{name} - {row['resume_name']}", content=html)


@app.delete("/api/candidates/<int:cid>/resume")
@login_required
def api_resume_delete(cid):
    row, err = _get_candidate_or_403(cid, need="delete")
    if err:
        return err
    if not row["resume_file"]:
        return jsonify({"error": "该候选人没有简历"}), 400
    _remove_resume_file(row["resume_file"])
    db = get_db()
    db.execute("UPDATE candidates SET resume_file=NULL, resume_name=NULL, updated_at=? WHERE id=?",
               (now_str(), cid))
    name = json.loads(row["data"]).get("name", "")
    add_log(g.user, "delete", f"{g.user['display_name']} 删除了「{name}」的简历（{row['resume_name']}）",
            cid, name, row["group_id"])
    db.commit()
    log.info("删除简历 %s cid=%d name=%s", _who(g.user), cid, name)
    return jsonify({"ok": True})


@app.post("/api/candidates/export")
@login_required
def api_candidates_export():
    """将勾选的候选人导出为 Excel：列 = 当前用户可见的字段配置 + 分组。"""
    ids = request.get_json(force=True).get("ids") or []
    if not ids:
        return jsonify({"error": "请先勾选候选人"}), 400
    db = get_db()
    placeholders = ",".join("?" * len(ids))
    rows = db.execute(f"SELECT * FROM candidates WHERE id IN ({placeholders})", ids).fetchall()

    group_id = None if g.user["role"] in GLOBAL_VIEW_ROLES else g.user["group_id"]
    fields = load_fields(group_id)
    names = group_name_map()

    wb = Workbook()
    ws = wb.active
    ws.title = "候选人导出"
    headers = ["分组"] + [f["label"] for f in fields]
    ws.append(headers)
    exported = 0
    for row in rows:
        if not can_view_group(g.user, row["group_id"]):
            continue
        data = json.loads(row["data"])
        ws.append([names.get(row["group_id"], "")] + [data.get(f["key"], "") for f in fields])
        exported += 1
    if exported == 0:
        return jsonify({"error": "选中的候选人均无导出权限"}), 403
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(12, len(h) * 2 + 4)

    add_log(g.user, "export", f"{g.user['display_name']} 导出了 {exported} 名候选人的Excel数据")
    db.commit()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    resp = send_file(buf, as_attachment=True, download_name=f"候选人导出_{ts}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp.headers["X-Export-Count"] = str(exported)
    return resp


@app.post("/api/resumes/export")
@login_required
def api_resumes_export():
    ids = request.get_json(force=True).get("ids") or []
    if not ids:
        return jsonify({"error": "请先勾选候选人"}), 400
    db = get_db()
    placeholders = ",".join("?" * len(ids))
    rows = db.execute(f"SELECT * FROM candidates WHERE id IN ({placeholders})", ids).fetchall()

    buf = io.BytesIO()
    exported, used_names = 0, set()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for row in rows:
            if not can_view_group(g.user, row["group_id"]) or not row["resume_file"]:
                continue
            path = os.path.join(RESUME_DIR, row["resume_file"])
            if not os.path.exists(path):
                continue
            name = json.loads(row["data"]).get("name", "未命名")
            arcname = f"{name}_{row['resume_name']}"
            if arcname in used_names:
                arcname = f"{name}_{row['id']}_{row['resume_name']}"
            used_names.add(arcname)
            zf.write(path, arcname)
            exported += 1
    if exported == 0:
        return jsonify({"error": "选中的候选人均没有可导出的简历"}), 400

    add_log(g.user, "export", f"{g.user['display_name']} 批量导出了 {exported} 份简历")
    db.commit()
    buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    resp = send_file(buf, as_attachment=True, download_name=f"简历导出_{ts}.zip",
                     mimetype="application/zip")
    resp.headers["X-Export-Count"] = str(exported)
    return resp


# ---------------------------------------------------------------- Excel 导入/模板
@app.get("/api/import/template")
@login_required
def api_template():
    fields = [f for f in load_fields() if f.get("importable")]
    wb = Workbook()
    ws = wb.active
    ws.title = "候选人导入模板"
    for i, f in enumerate(fields, start=1):
        ws.cell(row=1, column=i, value=f["excel_column"])
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(14, len(f["excel_column"]) * 2 + 4)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="候选人导入模板.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.post("/api/import")
@login_required
def api_import():
    if "file" not in request.files:
        return jsonify({"error": "请选择Excel文件"}), 400
    group_id = request.form.get("group_id", type=int) or g.user["group_id"]
    if not group_id:
        return jsonify({"error": "请指定导入的分组"}), 400
    if not can_edit_group(g.user, group_id):
        log.warning("Excel导入权限拒绝 %s group_id=%s", _who(g.user), group_id)
        return jsonify({"error": "无该分组的编辑权限"}), 403

    try:
        wb = load_workbook(request.files["file"], data_only=True)
    except Exception:
        return jsonify({"error": "文件解析失败，请上传 .xlsx 格式文件"}), 400
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({"error": "Excel内容为空"}), 400

    fields = [f for f in load_fields() if f.get("importable")]
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_map = {}  # 列下标 -> field key
    for idx, h in enumerate(header):
        for f in fields:
            if h == f["excel_column"] or h == f["label"]:
                col_map[idx] = f["key"]
                break
    if "name" not in col_map.values():
        return jsonify({"error": "Excel中未找到“候选人”列，请参考导入模板"}), 400
    log.debug("Excel导入列映射 %s group_id=%s cols=%s", _who(g.user), group_id, col_map)

    def cell_str(v):
        if v is None:
            return ""
        if isinstance(v, datetime):
            return v.strftime("%Y-%m-%d")
        return str(v).strip()

    db = get_db()
    existing = db.execute("SELECT * FROM candidates WHERE group_id=?", (group_id,)).fetchall()
    by_phone = {}
    by_name = {}
    for r in existing:
        d = json.loads(r["data"])
        if d.get("phone"):
            by_phone[d["phone"]] = r
        if d.get("name"):
            by_name[d["name"]] = r

    created = updated = skipped = 0
    for raw in rows[1:]:
        data = {}
        for idx, key in col_map.items():
            if idx < len(raw):
                data[key] = cell_str(raw[idx])
        if not data.get("name"):
            skipped += 1
            continue
        match = by_phone.get(data.get("phone", "")) or by_name.get(data["name"])
        if match:
            old = json.loads(match["data"])
            merged = dict(old)
            changed = False
            for k, v in data.items():
                if v and str(old.get(k, "")) != v:
                    merged[k] = v
                    changed = True
            if changed:
                db.execute("UPDATE candidates SET data=?, updated_at=? WHERE id=?",
                           (json.dumps(merged, ensure_ascii=False), now_str(), match["id"]))
                updated += 1
        else:
            cur = db.execute("INSERT INTO candidates (group_id, data, created_at, updated_at) VALUES (?,?,?,?)",
                             (group_id, json.dumps(data, ensure_ascii=False), now_str(), now_str()))
            by_name[data["name"]] = db.execute("SELECT * FROM candidates WHERE id=?", (cur.lastrowid,)).fetchone()
            created += 1

    gname = group_name_map().get(group_id, "")
    add_log(g.user, "import",
            f"{g.user['display_name']} 通过Excel向「{gname}」导入候选人：新增{created}人，更新{updated}人"
            + (f"，跳过{skipped}行" if skipped else ""),
            None, None, group_id)
    db.commit()
    log.info("Excel导入 %s group=%s 新增%d 更新%d 跳过%d",
             _who(g.user), gname, created, updated, skipped)
    return jsonify({"ok": True, "created": created, "updated": updated, "skipped": skipped})


# ---------------------------------------------------------------- 日志查询
@app.get("/api/logs")
@login_required
def api_logs():
    page = max(1, request.args.get("page", 1, type=int))
    size = APP_CONFIG["logs"]["page_size"]
    db = get_db()
    where, params = "", []
    if g.user["role"] not in GLOBAL_VIEW_ROLES:
        where = "WHERE group_id=? OR group_id IS NULL"
        params.append(g.user["group_id"] or -1)
    total = db.execute(f"SELECT COUNT(*) AS c FROM logs {where}", params).fetchone()["c"]
    rows = db.execute(
        f"SELECT * FROM logs {where} ORDER BY id DESC LIMIT ? OFFSET ?",
        params + [size, (page - 1) * size],
    ).fetchall()
    return jsonify({"total": total, "page": page, "size": size, "items": [dict(r) for r in rows]})


# ---------------------------------------------------------------- 全局总览（管理员/全局查看员）
@app.get("/api/overview")
@login_required
def api_overview():
    if g.user["role"] not in GLOBAL_VIEW_ROLES:
        return jsonify({"error": "需要管理员或全局查看员权限"}), 403
    db = get_db()
    names = group_name_map()
    groups = []
    for gid, gname in names.items():
        rows = db.execute("SELECT * FROM candidates WHERE group_id=? ORDER BY updated_at DESC", (gid,)).fetchall()
        cands = []
        stats = {"total": len(rows), "signed": 0, "onboarded": 0, "high_risk": 0}
        for r in rows:
            c = candidate_dict(r, names)
            log = db.execute(
                "SELECT message, created_at FROM logs WHERE candidate_id=? ORDER BY id DESC LIMIT 1", (r["id"],)
            ).fetchone()
            c["latest_log"] = (f"[{log['created_at']}] {log['message']}" if log else "暂无更新记录")
            cands.append(c)
            d = c["data"]
            if d.get("sign_status") == "已签约":
                stats["signed"] += 1
            if d.get("onboarded") == "是":
                stats["onboarded"] += 1
            if d.get("onboard_risk") == "高":
                stats["high_risk"] += 1
        groups.append({"group_id": gid, "group_name": gname, "stats": stats, "candidates": cands})
    return jsonify(groups)


# ---------------------------------------------------------------- 数据备份与恢复（仅管理员）
BACKUP_NAME_RE = re.compile(r"^backup_\d{8}_\d{6}(_\d{3})?(_manual)?\.db$")


def take_backup(manual=False):
    """用 SQLite 在线备份 API 生成数据库快照（WAL 模式下安全，不阻塞读写）。"""
    os.makedirs(BACKUP_DIR, exist_ok=True)
    # 文件名带毫秒，避免同一秒内连续备份（如恢复前的自动保存）互相覆盖
    ts = datetime.now()
    name = ("backup_" + ts.strftime("%Y%m%d_%H%M%S") + f"_{ts.microsecond // 1000:03d}"
            + ("_manual" if manual else "") + ".db")
    path = os.path.join(BACKUP_DIR, name)
    src = sqlite3.connect(DB_PATH)
    dst = sqlite3.connect(path)
    try:
        with dst:
            src.backup(dst)
    finally:
        src.close()
        dst.close()
    _prune_backups()
    log.debug("数据库备份完成 %s manual=%s", name, manual)
    return name


def _prune_backups():
    """按配置保留期清理过期快照（默认3天，每小时一份约72份）。"""
    retention_days = APP_CONFIG.get("backup", {}).get("retention_days", 3)
    cutoff = time.time() - retention_days * 86400
    for f in os.listdir(BACKUP_DIR):
        path = os.path.join(BACKUP_DIR, f)
        if BACKUP_NAME_RE.match(f) and os.path.getmtime(path) < cutoff:
            os.remove(path)


def backup_scheduler():
    """后台守护线程：启动时立即备份一次，之后按配置间隔（默认每小时）自动备份。"""
    interval = APP_CONFIG.get("backup", {}).get("interval_hours", 1) * 3600
    while True:
        try:
            name = take_backup()
            log.info("自动备份完成 %s", name)
        except Exception as e:
            log.error("自动备份失败: %s", e)
        time.sleep(interval)


@app.get("/api/backups")
@admin_required
def api_backups():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    items = []
    for f in os.listdir(BACKUP_DIR):
        if not BACKUP_NAME_RE.match(f):
            continue
        path = os.path.join(BACKUP_DIR, f)
        items.append({
            "name": f,
            "manual": "_manual" in f,
            "time": datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d %H:%M:%S"),
            "size_kb": round(os.path.getsize(path) / 1024, 1),
        })
    items.sort(key=lambda x: x["name"], reverse=True)
    return jsonify(items)


@app.post("/api/backups")
@admin_required
def api_backup_create():
    name = take_backup(manual=True)
    add_log(g.user, "backup", f"{g.user['display_name']} 手动创建了数据备份（{name}）")
    get_db().commit()
    log.info("手动备份 %s file=%s", _who(g.user), name)
    return jsonify({"ok": True, "name": name})


@app.post("/api/backups/restore")
@admin_required
def api_backup_restore():
    name = (request.get_json(force=True).get("name") or "").strip()
    if not BACKUP_NAME_RE.match(name):
        return jsonify({"error": "备份文件名不合法"}), 400
    path = os.path.join(BACKUP_DIR, name)
    if not os.path.exists(path):
        return jsonify({"error": "备份文件不存在"}), 404

    # 恢复前自动保存当前状态，误恢复可再恢复回来
    safety = take_backup(manual=True)

    src = sqlite3.connect(path)
    dst = sqlite3.connect(DB_PATH)
    try:
        with dst:
            src.backup(dst)
    finally:
        src.close()
        dst.close()

    # 恢复后的库写入本次操作日志
    add_log(g.user, "backup", f"{g.user['display_name']} 将数据恢复至备份「{name}」（恢复前状态已自动保存为 {safety}）")
    get_db().commit()
    log.info("恢复备份 %s target=%s safety=%s", _who(g.user), name, safety)
    return jsonify({"ok": True, "restored": name, "safety_backup": safety})


# ---------------------------------------------------------------- 静态页面
@app.get("/")
def index():
    return send_from_directory(app.static_folder, "index.html")


def port_in_use(port):
    """Windows 下多个进程可同时绑定同一端口导致请求被旧实例接管，启动前先探测。"""
    import socket
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.settimeout(1)
        return s.connect_ex(("127.0.0.1", port)) == 0


if __name__ == "__main__":
    init_db(demo="--demo" in sys.argv)
    # 端口/线程优先级：环境变量 > app_config.json
    port = int(os.environ.get("PORT", APP_CONFIG["server"]["port"]))
    threads = int(os.environ.get("THREADS", APP_CONFIG["server"]["threads"]))
    if port_in_use(port):
        log.error("端口 %d 已被占用，启动中止", port)
        print(f"错误：端口 {port} 已有服务在运行，请先停止旧实例（或修改 config/app_config.json 中的端口）。")
        sys.exit(1)
    # 启动定时备份线程（启动即备份一次，之后每 interval_hours 小时一份，保留 retention_days 天）
    threading.Thread(target=backup_scheduler, daemon=True).start()
    log.info("服务启动 port=%d threads=%d log_level=%s", port, threads,
             os.environ.get("LOG_LEVEL") or APP_CONFIG.get("logging", {}).get("level", "INFO"))
    print(f"校招入职跟踪管理系统已启动: http://127.0.0.1:{port} （waitress，{threads} 工作线程）")
    from waitress import serve
    serve(app, host="0.0.0.0", port=port, threads=threads,
          connection_limit=1024, channel_timeout=120)
